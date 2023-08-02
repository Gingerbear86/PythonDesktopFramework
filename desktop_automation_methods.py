"""
This module contains the code for housing my methods
that are used in my automated desktop testing.
"""

import os
import configparser
import logging
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from combined_automation_methods import CombinedAutomationMethods

logging.basicConfig(level=logging.INFO)


class DesktopAutomationMethods(CombinedAutomationMethods):
    """
    These are my desktop automation methods, and this is its class
    """
    def __init__(self):
        super().__init__()
        self.driver = None
        self.wait = None
        self.file_path = None

        config_handler = ConfigHandler()
        edge_driver_path = config_handler.get(
            'Selenium',
            'edge_driver_path'
        )
        edge_binary_path = config_handler.get(
            'Selenium',
            'edge_binary_path'
        )
        window_x_position = config_handler.getint(
           'Selenium',
           'window_position_x'
        )
        window_y_position = config_handler.getint(
           'Selenium',
           'window_position_y'
        )
        window_width = config_handler.getint(
           'Selenium',
           'window_width'
        )
        window_height = config_handler.getint(
           'Selenium',
           'window_height'
        )
        wait_timeout = config_handler.getint(
            'Selenium',
            'wait_timeout'
        )

        edge_options = webdriver.EdgeOptions()
        edge_options.add_argument("--log-level=3")
        edge_options.binary_location = edge_binary_path
        edge_options.add_argument('--inprivate')

        self.driver = webdriver.Edge(
            service=Service(
                executable_path=edge_driver_path, service_args=[
                    '--log-level=OFF'
                    ]
                ),
            options=edge_options
        )
        self.driver.set_window_position(window_x_position, window_y_position)
        self.driver.set_window_size(window_width, window_height)
        self.driver.maximize_window()

        self.wait = WebDriverWait(self.driver, wait_timeout)


class ConfigHandler:
    """
    ConfigHandler Class
    """
    def __init__(self, file_path='config.ini'):
        self.file_path = file_path
        self.config = configparser.ConfigParser()
        self._load_config()

    def _load_config(self):
        if os.path.exists(self.file_path):
            self.config.read(self.file_path)
        else:
            raise FileNotFoundError(f"{self.file_path} not found.")

    def get(self, section, option):
        """
        ConfigHandler Get
        """
        return self.config.get(section, option)

    def getint(self, section, option):
        """
        Get Int
        """
        return self.config.getint(section, option)


class WebElementHandler(DesktopAutomationMethods):
    """
    WebElementHandler Class
    """
    # Finds elements with waits and exception checking
    def find_element(self, locator):
        """
        Finds an element on the webpage
        Waits for the element to be clickable
        """
        try:
            self.wait.until(EC.presence_of_element_located(locator))
            logging.info("Element located: %s", locator)

            self.wait.until(EC.visibility_of_element_located(locator))
            logging.info("Element visible: %s", locator)

            self.wait.until(EC.element_to_be_clickable(locator))
            logging.info("Element clickable: %s", locator)

            element = self.driver.find_element(*locator)
            logging.info("Element found: %s, returning element.", locator)
            return element
        except Exception as exception:
            logging.error("Error finding element %s: %s", locator, str(
                exception
                ))
            raise

    # Finds elements then clicks them with a try, except loop.
    def find_element_and_click(self, locator):
        """
        Finds an element on the webpage, scrolls to it if needed, and clicks it
        """
        element = self.find_element(locator)
        logging.info("Element found.")
        try:
            element.click()
        except ElementClickInterceptedException:
            action = ActionChains(self.driver)
            action.move_to_element(element).click().perform()
        logging.info("Element clicked.")

    # Enters text into a field
    def enter_text(self, locator, text):
        """
        Enters text into a field on the webpage
        """
        element = self.find_element(locator)
        logging.info("Element found.")
        element.clear()
        logging.info("Element cleared.")
        element.send_keys(text)
        logging.info("Sending keys to element.")

    # Copies text from a field
    def copy_text(self, locator):
        """
        Copies text from a field on the webpage
        """
        element = self.find_element(locator)
        logging.info("Element found, returning element.")
        return element.text

    # Verifies the state of an element with built in wait and exceptions
    def verify_element_selection_state(self, locator):
        """
        Finds and reports the aria state of an element
        """
        self.wait.until(
            EC.visibility_of_element_located(locator)
        )
        self.wait.until(
            EC.element_to_be_clickable(locator)
        )
        element_state = self.driver.find_element(*locator)
        logging.info("Element found, returning state true or false.")
        return bool(element_state.get_attribute('aria-selected'))


class ExcelHandler(DesktopAutomationMethods):
    """
    ExcelHandler Class
    """
    # Writes table data to an excel file using Openpyxl
    def record_table_data(self, table_locator, file_path):
        """
        saves data from web tables to xlsx using openpyxl
        """
        table = self.driver.find_element(*table_locator)
        logging.info("Element from table found.")
        rows = table.find_elements(By.TAG_NAME, "tr")
        logging.info("Element from row found.")

        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            workbook.save(file_path)
            logging.info("Failed to find element, closing workbook.")
            workbook.close()

        try:
            workbook = openpyxl.load_workbook(file_path)
        except IOError:
            print(f"Could not open or find the file: {file_path}")
        worksheet = workbook.active
        logging.info("Workbook is now active.")

        header_cells = rows[0].find_elements(By.TAG_NAME, "th")
        logging.info("Element in row found.")
        for col_index, cell in enumerate(header_cells, start=1):
            cell_value = cell.text
            worksheet.cell(row=1, column=col_index, value=cell_value)
        logging.info("Elements in row copied to workbook.")

        # Loop through each row in the table
        for row_index, row in enumerate(rows, start=1):
            # Find each cell in the row
            cells = row.find_elements(By.TAG_NAME, "td")
            # Loop through each cell in the row
            for col_index, cell in enumerate(cells, start=1):
                # Extract the text value from the cell
                cell_value = cell.text
                # Write the cell value to the workbook
                logging.info(
                    "Elements in row's copied, writing to workbook."
                )
                worksheet.cell(
                    row=row_index,
                    column=col_index,
                    value=cell_value
                )

        workbook.save(file_path)
        logging.info("Workbook saved.")
        workbook.close()
        return file_path
